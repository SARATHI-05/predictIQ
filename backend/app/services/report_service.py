import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy.orm import Session
from app.models.food_record import FoodRecord
from app.models.prediction import Prediction
from app.models.inventory import InventoryItem
from app.models.audit_log import AuditLog
from app.models.prediction_accuracy import PredictionAccuracy
from typing import Dict, Any, List, Optional
from datetime import date

def generate_report_data(
    db: Session,
    report_type: str = "food_demand",
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    food_category: Optional[str] = None
) -> Dict[str, Any]:
    """Fetch filtered report items across all modules and compute summary metrics"""
    if report_type in ["food_demand", "wastage", "historical"]:
        query = db.query(FoodRecord)
        if start_date:
            query = query.filter(FoodRecord.date >= start_date)
        if end_date:
            query = query.filter(FoodRecord.date <= end_date)
        if food_category and food_category != "All":
            query = query.filter(FoodRecord.food_category == food_category)

        records = query.order_by(FoodRecord.date.desc()).all()

        total_prep = sum(r.food_prepared for r in records)
        total_cons = sum(r.food_consumed for r in records)
        total_left = sum(r.leftover for r in records)
        waste_pct = round((total_left / max(1, total_prep)) * 100, 2)

        items = [{
            'id': r.id,
            'date': r.date.strftime('%Y-%m-%d'),
            'food_category': r.food_category,
            'food_prepared': r.food_prepared,
            'food_consumed': r.food_consumed,
            'leftover': r.leftover,
            'expected_customers': r.expected_customers,
            'holiday': r.holiday,
            'special_event': r.special_event,
            'weather': r.weather,
            'wastage_percent': round((r.leftover / max(1, r.food_prepared)) * 100, 1)
        } for r in records]

        return {
            'report_type': report_type,
            'total_records': len(items),
            'summary': {
                'total_prepared': total_prep,
                'total_consumed': total_cons,
                'total_leftover': total_left,
                'wastage_percentage': waste_pct
            },
            'items': items
        }
    elif report_type == "inventory":
        query = db.query(InventoryItem)
        if food_category and food_category != "All":
            query = query.filter(InventoryItem.category == food_category)

        items = query.order_by(InventoryItem.category, InventoryItem.ingredient_name).all()
        total_val = sum(i.current_stock * i.unit_cost for i in items)
        low_stock = sum(1 for i in items if i.current_stock <= i.min_stock_level)

        item_list = [{
            'id': i.id,
            'ingredient_name': i.ingredient_name,
            'category': i.category,
            'current_stock': i.current_stock,
            'unit': i.unit,
            'min_stock_level': i.min_stock_level,
            'max_stock_level': i.max_stock_level,
            'unit_cost': i.unit_cost,
            'total_value': round(i.current_stock * i.unit_cost, 2),
            'supplier': i.supplier,
            'status': 'Low Stock' if i.current_stock <= i.min_stock_level else 'Stocked'
        } for i in items]

        return {
            'report_type': 'inventory',
            'total_records': len(item_list),
            'summary': {
                'total_items': len(item_list),
                'total_inventory_valuation': round(total_val, 2),
                'low_stock_items_count': low_stock
            },
            'items': item_list
        }
    elif report_type == "audit_logs":
        query = db.query(AuditLog)
        if start_date:
            query = query.filter(AuditLog.timestamp >= start_date)
        if end_date:
            query = query.filter(AuditLog.timestamp <= end_date)

        logs = query.order_by(AuditLog.id.desc()).limit(500).all()
        log_list = [{
            'id': l.id,
            'timestamp': l.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'user_email': l.user_email,
            'action': l.action,
            'module': l.module,
            'description': l.description,
            'ip_address': l.ip_address
        } for l in logs]

        return {
            'report_type': 'audit_logs',
            'total_records': len(log_list),
            'summary': {
                'total_logged_events': len(log_list),
                'unique_users': len(set(l['user_email'] for l in log_list if l['user_email']))
            },
            'items': log_list
        }
    elif report_type == "accuracy":
        query = db.query(PredictionAccuracy)
        if start_date:
            query = query.filter(PredictionAccuracy.prediction_date >= start_date)
        if end_date:
            query = query.filter(PredictionAccuracy.prediction_date <= end_date)
        if food_category and food_category != "All":
            query = query.filter(PredictionAccuracy.food_category == food_category)

        records = query.order_by(PredictionAccuracy.prediction_date.desc()).all()
        acc_list = [{
            'id': r.id,
            'date': r.prediction_date.strftime('%Y-%m-%d'),
            'food_category': r.food_category,
            'predicted_demand': r.predicted_demand,
            'actual_consumed': r.actual_consumed,
            'error': r.error,
            'abs_error': r.abs_error,
            'percentage_error': r.percentage_error,
            'accuracy_score': r.accuracy_score
        } for r in records]

        avg_acc = round(sum(r['accuracy_score'] for r in acc_list) / max(1, len(acc_list)), 2)
        avg_err = round(sum(r['abs_error'] for r in acc_list) / max(1, len(acc_list)), 1)

        return {
            'report_type': 'accuracy',
            'total_records': len(acc_list),
            'summary': {
                'total_evaluated_records': len(acc_list),
                'average_accuracy_percentage': avg_acc,
                'mean_absolute_error': avg_err
            },
            'items': acc_list
        }
    else:
        # Predictions report
        query = db.query(Prediction)
        if start_date:
            query = query.filter(Prediction.prediction_date >= start_date)
        if end_date:
            query = query.filter(Prediction.prediction_date <= end_date)
        if food_category and food_category != "All":
            query = query.filter(Prediction.food_category == food_category)

        predictions = query.order_by(Prediction.prediction_date.desc()).all()

        items = [{
            'id': p.id,
            'prediction_date': p.prediction_date.strftime('%Y-%m-%d'),
            'food_category': p.food_category,
            'expected_customers': p.expected_customers,
            'predicted_demand': p.predicted_demand,
            'recommended_preparation': p.recommended_preparation,
            'demand_level': p.demand_level,
            'model_version': p.model_version
        } for p in predictions]

        return {
            'report_type': 'predictions',
            'total_records': len(items),
            'summary': {
                'total_predictions': len(items),
                'avg_predicted_demand': round(sum(p['predicted_demand'] for p in items) / max(1, len(items)), 1)
            },
            'items': items
        }

def export_report_csv(data: Dict[str, Any]) -> str:
    """Generate CSV string from report data"""
    output = io.StringIO()
    items = data.get('items', [])
    if not items:
        return "No data available for the selected filters."

    writer = csv.DictWriter(output, fieldnames=list(items[0].keys()))
    writer.writeheader()
    writer.writerows(items)
    return output.getvalue()

def export_report_excel(data: Dict[str, Any]) -> bytes:
    """Generate styled Excel spreadsheet (.xlsx) in memory"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PredictIQ Report"

    # Header style
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center")

    items = data.get('items', [])
    if not items:
        ws.append(["No records found"])
    else:
        headers = [k.replace('_', ' ').title() for k in items[0].keys()]
        ws.append(headers)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center

        for row_idx, item in enumerate(items, start=2):
            row_values = list(item.values())
            ws.append(row_values)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def export_report_pdf_html(data: Dict[str, Any]) -> str:
    """Generate clean printable HTML format for PDF / browser printing"""
    report_title = data.get('report_type', 'Report').replace('_', ' ').title()
    summary = data.get('summary', {})
    items = data.get('items', [])

    html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<title>PredictIQ {report_title} Report</title>
<style>
  body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 30px; color: #1e293b; background: #fff; }}
  h1 {{ color: #0f172a; font-size: 24px; border-bottom: 2px solid #10b981; padding-bottom: 8px; }}
  .meta {{ color: #64748b; font-size: 12px; margin-bottom: 20px; }}
  .summary-box {{ display: flex; gap: 15px; margin-bottom: 25px; background: #f8fafc; padding: 15px; border-radius: 8px; border: 1px solid #e2e8f0; }}
  .kpi {{ flex: 1; }}
  .kpi-title {{ font-size: 11px; text-transform: uppercase; color: #64748b; }}
  .kpi-val {{ font-size: 18px; font-weight: bold; color: #0f172a; margin-top: 4px; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 15px; font-size: 12px; }}
  th, td {{ border: 1px solid #cbd5e1; padding: 8px 10px; text-align: left; }}
  th {{ background-color: #1e293b; color: #ffffff; font-weight: 600; }}
  tr:nth-child(even) {{ background-color: #f8fafc; }}
  .footer {{ margin-top: 30px; font-size: 11px; color: #94a3b8; text-align: center; }}
</style>
</head>
<body>
  <h1>PredictIQ &bull; {report_title} Report</h1>
  <div class="meta">Generated by PredictIQ AI Enterprise System &bull; Total Records: {len(items)}</div>
  
  <div class="summary-box">
"""
    for k, v in summary.items():
        k_fmt = k.replace('_', ' ').title()
        v_fmt = f"{v}%" if "percent" in k else (f"{v:,}" if isinstance(v, (int, float)) else str(v))
        html += f"""
    <div class="kpi">
      <div class="kpi-title">{k_fmt}</div>
      <div class="kpi-val">{v_fmt}</div>
    </div>"""

    html += """
  </div>
  <table>
    <thead><tr>"""

    if items:
        for k in items[0].keys():
            html += f"<th>{k.replace('_', ' ').upper()}</th>"
        html += "</tr></thead><tbody>"
        for row in items:
            html += "<tr>"
            for v in row.values():
                html += f"<td>{v}</td>"
            html += "</tr>"
    else:
        html += "<th>Notice</th></tr></thead><tbody><tr><td>No records found.</td></tr>"

    html += """
    </tbody>
  </table>
  <div class="footer">PredictIQ &bull; AI-Based Food Demand and Resource Planning System &bull; Confidential</div>
</body>
</html>"""
    return html

def bulk_delete_report_records(db: Session, report_type: str, ids: List[int], user_email: Optional[str] = None) -> int:
    """
    Perform bulk deletion on underlying database entities based on report module type.
    """
    if not ids:
        return 0

    deleted_count = 0
    from app.models.resource import ResourcePlan

    if report_type in ["food_demand", "wastage", "historical"]:
        records = db.query(FoodRecord).filter(FoodRecord.id.in_(ids)).all()
        deleted_count = len(records)
        for r in records:
            # Clean up linked accuracy record references if any
            db.query(PredictionAccuracy).filter(PredictionAccuracy.food_record_id == r.id).update(
                {"food_record_id": None}, synchronize_session=False
            )
            db.delete(r)

    elif report_type == "predictions":
        preds = db.query(Prediction).filter(Prediction.id.in_(ids)).all()
        deleted_count = len(preds)
        for p in preds:
            db.query(ResourcePlan).filter(ResourcePlan.prediction_id == p.id).delete(synchronize_session=False)
            db.query(PredictionAccuracy).filter(PredictionAccuracy.prediction_id == p.id).delete(synchronize_session=False)
            db.delete(p)

    elif report_type == "inventory":
        items = db.query(InventoryItem).filter(InventoryItem.id.in_(ids)).all()
        deleted_count = len(items)
        for item in items:
            db.delete(item)

    elif report_type == "audit_logs":
        logs = db.query(AuditLog).filter(AuditLog.id.in_(ids)).all()
        deleted_count = len(logs)
        for log in logs:
            db.delete(log)

    elif report_type == "accuracy":
        accs = db.query(PredictionAccuracy).filter(PredictionAccuracy.id.in_(ids)).all()
        deleted_count = len(accs)
        for a in accs:
            db.delete(a)

    else:
        # Default fallback to FoodRecord
        records = db.query(FoodRecord).filter(FoodRecord.id.in_(ids)).all()
        deleted_count = len(records)
        for r in records:
            db.delete(r)

    # Log audit entry
    if deleted_count > 0:
        audit_entry = AuditLog(
            user_email=user_email or "System",
            action="BULK_DELETE",
            module=f"Reports ({report_type})",
            record_id=f"Count: {deleted_count}",
            description=f"Bulk deleted {deleted_count} items from {report_type} report view (IDs: {ids[:15]}{'...' if len(ids)>15 else ''})"
        )
        db.add(audit_entry)

    db.commit()
    return deleted_count

